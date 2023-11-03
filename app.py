from flask import Flask, render_template, request, redirect, url_for, send_from_directory
from werkzeug.utils import secure_filename
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from database.conection import DatabaseConnection
from datetime import datetime

connection = DatabaseConnection()


def readDNIInData(entrada):
    try:
        dni_list = []
        wb = openpyxl.load_workbook(entrada)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            dni = row[2]
            if dni:
                dni_list.append(f"'{dni}'")
        resultado = ',\n'.join(dni_list)
        return resultado
    except FileNotFoundError:
        print("El archivo no fue encontrado.")
    except Exception as e:
        print(f"Error al leer el archivo: {e}")

def dnis_query(dnis):
    return f"""
        select pnet_DocumentNumber as DNI 
            from CRM.crm.LeadBase 
            where pnet_DocumentNumber in ({dnis})
        union all
        select pnet_DocumentNumber as DNI 
            from CRM.crm.ContactBase 
            where pnet_DocumentNumber in ({dnis})
        """
        
def zonification_query(localidad):
    return f"""
        select TOP (1) * from comercial.dbo.Zonificacion_IC where localidad like '%{localidad}%'
        """

def crear_archivo_sin_coincidencias(db_connection, entrada, archivo_salida):
    conexion = db_connection.connect()
    try:
        with conexion.cursor() as cursor:
            cursor.execute(dnis_query(readDNIInData(entrada)))
            results = cursor.fetchall()
            dnism = [str(row[0]) for row in results]
    except Exception as e:
        print(f'\nError al realizar la consulta: {e}\n')
        return
    finally:
        conexion.close()
    wb = openpyxl.load_workbook(entrada)
    sheet = wb.active
    wb_output = Workbook()
    sheet_output = wb_output.active
    sheet_output.append(next(sheet.iter_rows(values_only=True)))
    for row in sheet.iter_rows(min_row=2, values_only=True):
        dni = row[2]
        if dni:
            dni = str(dni)
            if dni not in dnism:
                sheet_output.append(row)
    wb_output.save(archivo_salida)
    print(f"\nSe han filtrado los datos correctamente.\n")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def get_months_diff(start_date_str):
    try:
        start_date = datetime.strptime(start_date_str, '%d/%m/%Y')
    except ValueError:
        start_date = datetime(2023, 1, 1, 0, 0)
    current_date = datetime.now()
    diff = current_date.year * 12 + current_date.month - (start_date.year * 12 + start_date.month)
    return max(diff, 1)

def convert_excel_to_txt(db_connection, excel_file, txt_file, como_se_entero):
    conexion = db_connection.connect()
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    with open(txt_file, 'w', encoding='utf-8') as txt_file:
        txt_file.write('tipo;dni;sexo;sucursal;nombre;apellido;rating;comoseentero;recomendadoportipo;recomendadopordni;recomendadoporsexo;codigodearea1;telefonotipo1;telefono1;codigodearea2;telefonotipo2;telefono2;codigodearea3;telefonotipo3;telefono3;email;microempresa;inicioactividades;permanencia;cnae;cno;segmento;actividad;categoria;accioncomercial;descripcionaccion;tipodoctributario;doctributariopart1;doctributariopart2;doctributariopart3;tipodireccion;calle;callenro;piso;departamento;distrito;localidad;codigopostal;provincia;pais\n')
        for row in sheet.iter_rows(values_only=True, min_row=2):
            ubicaciondelamicroempresa = row[11]
            tipo = 'DNI'
            dni = str(row[2])
            sexo = 'M' if row[3] == 'Hombre' else 'F'
            sucursal = ''
            if ubicaciondelamicroempresa == 'Provincia de Buenos Aires':
                with conexion.cursor() as cursor:
                    cursor.execute(zonification_query(row[12].strip())) 
                    result = cursor.fetchone()
                    if result:
                        sucursal = result[3]
                    else:
                        sucursal = 'Casa Matriz'
            elif ubicaciondelamicroempresa == 'CABA':
                if len(row) > 11 and row[13] is not None and row[13].strip():
                    with conexion.cursor() as cursor:
                        cursor.execute(zonification_query(row[13].strip()))
                        result = cursor.fetchone()
                        if result:
                            sucursal = result[3]                                    
                        else:
                            sucursal = 'Casa Matriz'
            nombre = row[0]
            apellido = row[1]  
            rating = '1'     
            if como_se_entero == '000':
                comoseentero = '>>>XXXXXX - NO TIENE ORIGEN DE DEMANDA - XXXXXX<<<'
            elif como_se_entero == '00':
                comoseentero = '100000000'
            elif como_se_entero == '01':
                comoseentero = '102610026'
            elif como_se_entero == '02':
                comoseentero = '102610015'
            elif como_se_entero == '03':
                comoseentero = '102610019'
            elif como_se_entero == '04':
                comoseentero = '102610000'
            elif como_se_entero == '05':
                comoseentero = '102610001'
            elif como_se_entero == '06':
                comoseentero = '102610002'
            elif como_se_entero == '07':
                comoseentero = '102610022'
            elif como_se_entero == '08':
                comoseentero = '102610003'
            elif como_se_entero == '09':
                comoseentero = '102610023'
            elif como_se_entero == '10':
                comoseentero = '102610024'
            elif como_se_entero == '11':
                comoseentero = '102610025'
            elif como_se_entero == '12':
                comoseentero = '102610004'
            elif como_se_entero == '13':
                comoseentero = '102610005'
            elif como_se_entero == '14':
                comoseentero = '102610014'
            elif como_se_entero == '15':
                comoseentero = '102610006'
            elif como_se_entero == '16':
                comoseentero = '102610018'
            elif como_se_entero == '17':
                comoseentero = '102610007'
            elif como_se_entero == '18':
                comoseentero = '102610008'
            elif como_se_entero == '19':
                comoseentero = '102610009'
            elif como_se_entero == '20':
                comoseentero = '102610010'
            elif como_se_entero == '21':
                comoseentero = '102610011'
            elif como_se_entero == '22':
                comoseentero = '102610012'
            elif como_se_entero == '23':
                comoseentero = '102610016'
            elif como_se_entero == '24':
                comoseentero = '102610017'
            elif como_se_entero == '25':
                comoseentero = '738920000'
            elif como_se_entero == '26':
                comoseentero = '102610013'
            elif como_se_entero == '27':
                comoseentero = '102610020'
            elif como_se_entero == '28':
                comoseentero = '102610021'
            else:
                comoseentero = ''            
            recomendadoportipo = ''
            recomendadopordni = ''
            recomendadoporsexo = ''
            codigodearea1 = str(row[5])
            telefonotipo1 = '102610001' if row[4] == 'Celular' else '102610000'
            telefono1 = str(row[6])
            if len(telefono1) == 10:
                telefono1 = "15" + telefono1[2:]
            elif len(telefono1) < 10:
                telefono1 = "15" + telefono1
            codigodearea2 = ''
            telefonotipo2 = ''
            telefono2 = ''
            codigodearea3 = ''
            telefonotipo3 = ''
            telefono3 = ''
            email = row[7]
            microempresa = ''            
            fecha_inicio = row[10]
            permanencia = row[10]
            if fecha_inicio is not None and isinstance(fecha_inicio, datetime):
                formatted_date = fecha_inicio.strftime('%Y-%m-%d 00:00:00')
                start_date_str = fecha_inicio.strftime('%d/%m/%Y')
                months_diff = get_months_diff(start_date_str)
                permanencia = str(months_diff)
            else:
                formatted_date = '2023-01-01 00:00:00'                
            cnae = ''            
            cno = ''            
            segmento = ''            
            actividad = ''            
            categoria = ''            
            accioncomercial = '1'         
            descripcionaccion = 'Importacion Manual desde la app'         
            tipodoctributario = ''            
            doctributariopart1 = ''            
            doctributariopart2 = ''            
            doctributariopart3 = ''            
            tipodireccion = '102610000'            
            calle = ''
            callenro = ''
            piso = ''
            departamento = ''
            distrito = ''
            if ubicaciondelamicroempresa == 'Provincia de Buenos Aires':
                with conexion.cursor() as cursor:
                    cursor.execute(zonification_query(row[12].strip())) 
                    result = cursor.fetchone()
                    if result:
                        distrito = result[0]
                    else:
                        distrito = 'CABA'
            elif ubicaciondelamicroempresa == 'CABA':
                if len(row) > 11 and row[13] is not None and row[13].strip():
                    with conexion.cursor() as cursor:
                        cursor.execute(zonification_query(row[13].strip()))
                        result = cursor.fetchone()
                        if result:
                            distrito = result[0]                                    
                        else:
                            distrito = 'CABA'
            localidad = ''
            if ubicaciondelamicroempresa == 'Provincia de Buenos Aires':
                with conexion.cursor() as cursor:
                    cursor.execute(zonification_query(row[12].strip())) 
                    result = cursor.fetchone()
                    if result:
                        localidad = result[1]
                    else:
                        localidad = 'Barrio de San Telmo'
            elif ubicaciondelamicroempresa == 'CABA':
                if len(row) > 11 and row[13] is not None and row[13].strip():
                    with conexion.cursor() as cursor:
                        cursor.execute(zonification_query(row[13].strip()))
                        result = cursor.fetchone()
                        if result:
                            localidad = result[1]                                    
                        else:
                            localidad = 'Barrio de San Telmo'
            codigopostal = ''
            provincia = '102610001'
            pais = ''   
            
            row_data = [
                    tipo, 
                    dni, 
                    sexo, 
                    sucursal, 
                    nombre, 
                    apellido, 
                    rating, 
                    comoseentero, 
                    recomendadoportipo, 
                    recomendadopordni, 
                    recomendadoporsexo, 
                    codigodearea1, telefonotipo1, 
                    telefono1, 
                    codigodearea2, 
                    telefonotipo2, 
                    telefono2, 
                    codigodearea3, 
                    telefonotipo3, 
                    telefono3, 
                    email, 
                    microempresa, 
                    formatted_date, 
                    permanencia,
                    cnae,           
                    cno,           
                    segmento,           
                    actividad,           
                    categoria,           
                    accioncomercial,         
                    descripcionaccion,         
                    tipodoctributario,           
                    doctributariopart1,           
                    doctributariopart2,           
                    doctributariopart3,           
                    tipodireccion,            
                    calle,
                    callenro,
                    piso,
                    departamento,
                    distrito,
                    localidad,
                    codigopostal,
                    provincia,
                    pais
                    ]
            txt_file.write(';'.join(row_data) + '\n')



app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']
        como_se_entero = request.form.get('comoseentero')
        if file.filename == '':
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            new_filename = f'salida.xlsx'
            new_file_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
            file.save(new_file_path)
            txt_filename = f'LEAD_IMPORT_MANUAL_{timestamp}.txt'
            txt_file_path = os.path.join(app.config['UPLOAD_FOLDER'], txt_filename)
            salida_filename = 'salida.xlsx'
            salida_file_path = os.path.join(app.config['UPLOAD_FOLDER'], salida_filename)
            crear_archivo_sin_coincidencias(connection, new_file_path, salida_file_path)
            convert_excel_to_txt(connection, new_file_path, txt_file_path, como_se_entero)
            return render_template('success.html', salida_filename=salida_filename)
    return render_template('upload.html')


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        return 'El archivo no existe.'
    txt_filename = 'LEAD_IMPORT.txt'
    txt_file_path = os.path.join(app.config['UPLOAD_FOLDER'], txt_filename)    
    print("file_path:", file_path)
    print("txt_file_path:", txt_file_path)    
    convert_excel_to_txt(connection, file_path, txt_file_path)    
    return redirect(url_for('upload_success'))

            
@app.route('/success')
def upload_success():
    return render_template('success.html')


@app.route('/download')
def download_file():
    txt_filename = f'LEAD_IMPORT_MANUAL_{datetime.now().strftime("%Y%m%d_%H%M")}.txt'
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], txt_filename)
    if os.path.exists(file_path):
        return send_from_directory(app.config['UPLOAD_FOLDER'], txt_filename, as_attachment=True)
    else:
        return f'El archivo {txt_filename} no existe.'


if __name__ == '__main__':
    app.run(debug=True)
    
    #prueba
    
