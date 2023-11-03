import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from database.conection import DatabaseConnection
from datetime import datetime

connection = DatabaseConnection()


def readDNIInData(entrada):
    try:
        dni_list = []
        wb = openpyxl.load_workbook(entrada)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            dni = row[2]
            if dni:
                dni_list.append(f"'{dni}'")
        resultado = ',\n'.join(dni_list)
        return resultado
    except FileNotFoundError:
        print("El archivo no fue encontrado.")
    except Exception as e:
        print(f"Error al leer el archivo: {e}")

def dnis_query(dnis):
    return f"""
        select pnet_DocumentNumber as DNI 
            from CRM.crm.LeadBase 
            where pnet_DocumentNumber in ({dnis})
        union all
        select pnet_DocumentNumber as DNI 
            from CRM.crm.ContactBase 
            where pnet_DocumentNumber in ({dnis})
        """
        
def zonification_query(localidad):
    return f"""
        select TOP (1) * from comercial.dbo.Zonificacion_IC where localidad like '%{localidad}%'
        """

def crear_archivo_sin_coincidencias(db_connection, entrada, archivo_salida):
    conexion = db_connection.connect()
    try:
        with conexion.cursor() as cursor:
            cursor.execute(dnis_query(readDNIInData(entrada)))
            results = cursor.fetchall()
            dnism = [str(row[0]) for row in results]
    except Exception as e:
        print(f'\nError al realizar la consulta: {e}\n')
        return
    finally:
        conexion.close()
    wb = openpyxl.load_workbook(entrada)
    sheet = wb.active
    wb_output = Workbook()
    sheet_output = wb_output.active
    sheet_output.append(next(sheet.iter_rows(values_only=True)))
    for row in sheet.iter_rows(min_row=2, values_only=True):
        dni = row[2]
        if dni:
            dni = str(dni)
            if dni not in dnism:
                sheet_output.append(row)
    wb_output.save(archivo_salida)
    print(f"\nSe han filtrado los datos correctamente.\n")


entrada = 'uploads/entrada.xlsx'
archivo_salida = 'prueba_salida_tst.xlsx'
crear_archivo_sin_coincidencias(connection, entrada, archivo_salida)


